from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from datetime import date
from django.utils import timezone
import json
from .models import Site, Department, DocumentTemplate, DailyReport, ReportVersion
from django.template import engines
from records.utils import get_active_site, get_allowed_sites  # Added get_allowed_sites for security checks
import openpyxl
from django.views.decorators.http import require_POST
from django.http import HttpResponse
# --- STEP 1: Dashboard (Departments Hub) ---
@login_required
def dashboard(request):
    active_site = get_active_site(request)
    
    # Filter departments strictly by the active site
    if active_site:
        departments = Department.objects.filter(site=active_site).prefetch_related('templates')
    else:
        departments = Department.objects.none()
        
    return render(request, 'records/dashboard.html', {'departments': departments})

# --- STEP 2: Templates in a Department ---
@login_required
def dept_templates(request, dept_id):
    # Security: Ensure the requested department belongs to an allowed site
    allowed_sites = get_allowed_sites(request)
    department = get_object_or_404(Department, id=dept_id, site__in=allowed_sites)
    
    # Only show active templates
    templates = department.templates.filter(is_active=True).order_by('title')
    return render(request, 'records/dept_templates.html', {
        'department': department, 
        'templates': templates
    })

# --- STEP 3: Record List & Toggles ---
@login_required
def template_records(request, template_id):
    # Security: Ensure the template's department belongs to an allowed site
    allowed_sites = get_allowed_sites(request)
    template = get_object_or_404(DocumentTemplate, id=template_id, department__site__in=allowed_sites)
    
    # Fetch all reports, newest first
    reports = DailyReport.objects.filter(template=template).order_by('-report_date', '-created_at')
    
    return render(request, 'records/template_records.html', {
        'template': template,
        'reports': reports,
        'has_custom_categories': template.tracking_type == 'TIME_AND_CUSTOM'
    })

# --- STEP 4: Record View with Version Sidebar & Signatures ---
@login_required
def record_view(request, report_id):
    allowed_sites = get_allowed_sites(request)
    report = get_object_or_404(DailyReport, id=report_id, template__department__site__in=allowed_sites)
    
    # Pre-fetch user features for RBAC checks
    user_features = []
    if hasattr(request.user, 'nexus_profile'):
        user_features = list(request.user.nexus_profile.roles.values_list('allowed_features__code_name', flat=True))
    
    if request.method == 'POST':
        data = json.loads(request.body)
        action = data.get('action', 'EDIT')
        
        # ---------------------------------------------------------
        # ACTION: DIGITAL SIGNATURE (No new version spawned)
        # ---------------------------------------------------------
        if action == 'SIGN':
            if request.user.is_superuser:
                return JsonResponse({'status': 'error', 'message': 'Superusers cannot sign compliance records.'}, status=403)

            level = data.get('level')
            
            if level == 'L1':
                req_feature = f"sign_l1_{report.template.l1_signature_role.lower()}"
                if req_feature not in user_features:
                    return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)
                
                report.l1_signed_by = request.user
                report.l1_signed_at = timezone.now()
                report.status = 'PENDING_L2' if report.template.l2_signature_role else 'LOCKED'
                report.save()
                return JsonResponse({'status': 'success'})

            elif level == 'L2':
                req_feature = f"sign_l2_{report.template.l2_signature_role.lower()}"
                if req_feature not in user_features:
                    return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)
                
                if not report.l1_signed_by:
                    return JsonResponse({'status': 'error', 'message': 'L1 Maker must sign first.'}, status=400)

                report.l2_signed_by = request.user
                report.l2_signed_at = timezone.now()
                report.status = 'LOCKED'
                report.save()
                return JsonResponse({'status': 'success'})

        # ---------------------------------------------------------
        # ACTION: MANUAL EDIT (Resets signatures to DRAFT)
        # ---------------------------------------------------------
        elif action == 'EDIT':
            updated_data = data.get('updated_data', {})
            update_reason = data.get('update_reason', 'Manual Data Correction')
            
            # Update data
            report.latest_data = updated_data
            
            # ISO COMPLIANCE: Editing a document resets the signature flow
            report.l1_signed_by = None
            report.l1_signed_at = None
            report.l2_signed_by = None
            report.l2_signed_at = None
            report.status = 'DRAFT'
            report.save()
            
            version_num = report.versions.count() + 1
            ReportVersion.objects.create(
                daily_report=report,
                version_number=version_num,
                action_type='MANUAL_EDIT',
                update_reason=update_reason,
                data_snapshot=updated_data,
                modified_by=request.user if request.user.is_authenticated else None
            )
            return JsonResponse({'status': 'success'})

    # --- GET REQUEST: ASSEMBLE UI STATE ---
    can_sign_l1 = False
    can_sign_l2 = False

    # Superusers NEVER get signing buttons
    if not request.user.is_superuser:
        if report.template.l1_signature_role:
            req_l1 = f"sign_l1_{report.template.l1_signature_role.lower()}"
            can_sign_l1 = req_l1 in user_features

        if report.template.l2_signature_role:
            req_l2 = f"sign_l2_{report.template.l2_signature_role.lower()}"
            can_sign_l2 = req_l2 in user_features

    versions = report.versions.all().order_by('-version_number')
    versions_data_list = [{'v_num': v.version_number, 'data': v.data_snapshot} for v in versions]
    
    return render(request, 'records/record_view.html', {
        'report': report,
        'template': report.template,
        'versions': versions,
        'versions_json': json.dumps(versions_data_list),
        'can_sign_l1': can_sign_l1,
        'can_sign_l2': can_sign_l2,
    })

# --- TEMPLATE BUILDER ---
@login_required
def template_builder(request):
    active_site = get_active_site(request)
    
    if request.method == 'POST':
        title = request.POST.get('title')
        format_id = request.POST.get('format_id')
        dept_id = request.POST.get('department_id')
        doc_type = request.POST.get('document_type')
        frequency = request.POST.get('frequency')
        tracking_type = request.POST.get('tracking_type')
        html_layout = request.POST.get('html_layout')
        schema = request.POST.get('ai_prompt_schema')
        custom_extraction_prompt = request.POST.get('custom_extraction_prompt')
        
        # Capture the Compliance Signature Roles
        l1_role = request.POST.get('l1_signature_role')
        l2_role = request.POST.get('l2_signature_role')
        
        # Convert empty HTML strings to None for clean DB storage
        if not l1_role: l1_role = None
        if not l2_role: l2_role = None
        
        # Safely parse the hidden custom categories JSON string from the frontend
        custom_categories_raw = request.POST.get('custom_categories', '[]')
        try:
            custom_categories = json.loads(custom_categories_raw)
        except json.JSONDecodeError:
            custom_categories = []
        
        # Security: Ensure they can only create templates for their allowed sites
        allowed_sites = get_allowed_sites(request)
        department = get_object_or_404(Department, id=dept_id, site__in=allowed_sites)
        
        parsed_schema = None
        if doc_type != 'PRINT_ONLY' and schema and schema.strip():
            try:
                parsed_schema = json.loads(schema)
            except json.JSONDecodeError:
                messages.error(request, "Invalid JSON Schema provided.")
                return redirect('records:template_builder')

        # Create the actual template
        template = DocumentTemplate.objects.create(
            title=title,
            format_id=format_id,
            department=department,
            document_type=doc_type,
            frequency=frequency,
            tracking_type=tracking_type,
            custom_categories=custom_categories,
            html_layout=html_layout,
            ai_prompt_schema=parsed_schema,
            custom_extraction_prompt=custom_extraction_prompt,
            l1_signature_role=l1_role,
            l2_signature_role=l2_role
        )
        
        # --- NEW: AUTO-CREATE ANCHOR RECORD & VERSION 1 FOR STATIC DOCUMENTS ---
        if doc_type == 'PRINT_ONLY':
            report = DailyReport.objects.create(
                template=template,
                report_date=date.today(),
                latest_data={}, # Static docs don't need extracted JSON data
                status='DRAFT'  # Ready for the Maker to sign!
            )
            
            # Explicitly create the first audit trail entry and force it to Version 1
            ReportVersion.objects.create(
                daily_report=report,
                version_number=1,  # <--- Guarantees it starts at v1
                action_type='MANUAL_EDIT',
                update_reason="Initial Static Document Deployed",
                data_snapshot={},
                modified_by=request.user
            )
        
        messages.success(request, f"Template '{title}' created successfully!")
        return redirect('records:dashboard')
        
    # GET Request: Filter the dropdown so they can only build templates for the active site
    if active_site:
        departments = Department.objects.filter(site=active_site)
    else:
        departments = Department.objects.none()
        
    return render(request, 'records/template_builder.html', {'departments': departments})

    # --- SCAN/UPLOAD VIEW ---
@login_required
def scan_record(request, template_id):
    # Security: Ensure the template belongs to an allowed site
    allowed_sites = get_allowed_sites(request)
    template = get_object_or_404(DocumentTemplate, id=template_id, department__site__in=allowed_sites)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        extracted_json = data.get('extracted_data', {})
        
        manual_date = data.get('manual_date')
        update_reason = data.get('update_reason', 'Initial Record Entry') 
        
        # 1. Dynamic Anchor Extraction (Prioritize Manual Date)
        if manual_date:
            report_date = manual_date
        elif extracted_json.get('nexus_global_date'):
            report_date = extracted_json.get('nexus_global_date')
        else:
            report_date = date.today().isoformat()
            
        shift = extracted_json.get('val_shift', None) 
        
        # 2. Machine ID / Custom Parameter extraction
        machine_id = None
        if template.tracking_type == 'TIME_AND_CUSTOM':
            # NEW: Prioritize the explicit UI dropdown selection over the AI's guess
            ui_category = data.get('custom_category')
            if ui_category:
                machine_id = ui_category
            else:
                machine_id = extracted_json.get('nexus_global_machine', 'Unspecified')
        
        # 3. Get or Create the Parent DailyReport
        report, created = DailyReport.objects.get_or_create(
            template=template,
            report_date=report_date,
            shift=shift,
            machine_id=machine_id,
            defaults={'latest_data': extracted_json}
        )
        
        # It is ALWAYS a scan upload if it comes from this view!
        action_type = 'NEW_UPLOAD' 
        
        if not created:
            # Overwrite the latest data for reporting purposes
            report.latest_data = extracted_json
            report.save()
            
        # 4. Create the Child ReportVersion (Audit Trail)
        version_num = report.versions.count() + 1
        ReportVersion.objects.create(
            daily_report=report,
            version_number=version_num,
            action_type=action_type,
            update_reason=update_reason,
            data_snapshot=extracted_json,
            # scanned_file = will be handled later when we add S3 file uploading
            modified_by=request.user if request.user.is_authenticated else None
        )
        
        return JsonResponse({'status': 'success', 'message': f'Record saved (Version {version_num}).'})
        
    # Grab all existing record dates so the frontend knows when to show the "Update Reason" box
    existing_reports = list(template.daily_reports.values('report_date'))
    existing_dates = [str(r['report_date']) for r in existing_reports]
        
    return render(request, 'records/scan_record.html', {
        'template': template,
        'existing_dates_json': json.dumps(existing_dates)
    })

@require_POST
@login_required
def api_excel_to_html(request):
    """
    Takes an uploaded .xlsx file and mathematically maps its grid to an HTML table.
    Focuses purely on structural and visual translation.
    """
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active

        html_lines = ['<table class="nexus-grid" style="width: 100%; border-collapse: collapse;">', '  <tbody>']

        merged_cells = sheet.merged_cells.ranges
        skip_cells = set()
        merge_info = {}

        # 1. Map merged cells
        for merged_range in merged_cells:
            min_col, min_row, max_col, max_row = merged_range.bounds
            colspan = max_col - min_col + 1
            rowspan = max_row - min_row + 1
            top_left = (min_row, min_col)
            merge_info[top_left] = {'colspan': colspan, 'rowspan': rowspan}

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) != top_left:
                        skip_cells.add((r, c))

        max_r = sheet.max_row
        max_c = sheet.max_column

        # 2. Build the HTML Grid
        for r in range(1, max_r + 1):
            html_lines.append('    <tr>')
            for c in range(1, max_c + 1):
                if (r, c) in skip_cells:
                    continue

                td_attrs = []
                if (r, c) in merge_info:
                    colspan = merge_info[(r, c)]['colspan']
                    rowspan = merge_info[(r, c)]['rowspan']
                    if colspan > 1: td_attrs.append(f'colspan="{colspan}"')
                    if rowspan > 1: td_attrs.append(f'rowspan="{rowspan}"')

                cell = sheet.cell(row=r, column=c)
                cell_val = cell.value
                style_str = ""

                # Extract Formatting
                if cell.font and cell.font.bold:
                    style_str += "font-weight: bold; "
                if cell.alignment:
                    if cell.alignment.horizontal: style_str += f"text-align: {cell.alignment.horizontal}; "
                    if cell.alignment.vertical: style_str += f"vertical-align: {cell.alignment.vertical}; "
                
                # Extract Background Colors Safely
                if cell.fill and cell.fill.start_color:
                    color_val = cell.fill.start_color.rgb
                    if color_val and isinstance(color_val, str) and color_val != '00000000': 
                        if len(color_val) == 8: 
                            hex_color = '#' + color_val[2:]
                        else: 
                            hex_color = '#' + color_val
                        style_str += f"background-color: {hex_color} !important; "
                
                if style_str:
                    td_attrs.append(f'style="{style_str.strip()}"')

                # Content Logic & Variable Extraction
                val_str = str(cell_val).strip() if cell_val is not None else ""

                if val_str.startswith("logo-name="):
                    # DYNAMIC LOGO INJECTION
                    logo_filename = val_str.replace("logo-name=", "").strip()
                    # Using a flexbox wrapper ensures absolute dead-center alignment
                    content = f'<div style="display: flex; justify-content: center; align-items: center; width: 100%; height: 100%;"><img src="/media/logos/{logo_filename}" style="max-width: 150px; object-fit: contain;"></div>'
                
                elif val_str == "---":
                    # STATIC BLANK CONVENTION
                    content = "&nbsp;"
                
                elif val_str != "":
                    # STANDARD TEXT
                    content = val_str.replace('\n', '<br>')
                
                else:
                    # EDITABLE INPUT
                    cell_id = f"field_r{r}_c{c}"
                    content = f'<span contenteditable="true" class="editable-cell" id="{cell_id}"></span>'

                attr_str = " ".join(td_attrs)
                html_lines.append(f'      <td {attr_str}>{content}</td>')
            html_lines.append('    </tr>')

        html_lines.append('  </tbody>\n</table>')

        return JsonResponse({
            'html': "\n".join(html_lines)
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Separated correctly with line breaks!
@login_required
def excel_to_html(request):
    return render(request, 'records/excel_to_html.html')

# --- TEMPLATE SYNC HUB ---
@login_required
def template_sync_view(request):
    """ Renders the UI for importing and exporting templates """
    return render(request, 'records/template_sync.html')

@login_required
def export_templates(request):
    """ Serializes all allowed templates into a downloadable JSON file """
    allowed_sites = get_allowed_sites(request)
    templates = DocumentTemplate.objects.filter(department__site__in=allowed_sites)
    
    export_data = []
    for t in templates:
        export_data.append({
            'title': t.title,
            'format_id': t.format_id,
            'department_name': t.department.name,
            'site_name': t.department.site.name,
            'document_type': t.document_type,
            'frequency': t.frequency,
            'tracking_type': t.tracking_type,
            'custom_categories': t.custom_categories,
            'html_layout': t.html_layout,
            'ai_prompt_schema': t.ai_prompt_schema,
            'custom_extraction_prompt': t.custom_extraction_prompt,
            'l1_signature_role': t.l1_signature_role,
            'l2_signature_role': t.l2_signature_role,
        })
    
    response = HttpResponse(json.dumps(export_data, indent=4), content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="nexus_templates_export.json"'
    return response

@login_required
def import_templates(request):
    """ Reads a JSON file and injects templates, auto-mapping departments """
    if request.method == 'POST' and request.FILES.get('import_file'):
        import_file = request.FILES['import_file']
        try:
            data = json.loads(import_file.read().decode('utf-8'))
            imported_count = 0
            
            for item in data:
                site_name = item.get('site_name')
                dept_name = item.get('department_name')
                title = item.get('title')
                doc_type = item.get('document_type', 'PHYSICAL_OCR')
                
                # SAFETY: Skip malformed JSON blocks
                if not site_name or not dept_name or not title:
                    continue
                
                # 1. Resolve Site & Department
                site, _ = Site.objects.get_or_create(name=site_name)
                department, _ = Department.objects.get_or_create(name=dept_name, site=site)
                
                # 2. Create or Update Template safely
                template, created = DocumentTemplate.objects.update_or_create(
                    title=title,
                    department=department,
                    defaults={
                        'format_id': item.get('format_id'),
                        'document_type': doc_type,
                        'frequency': item.get('frequency', 'DAILY'),
                        'tracking_type': item.get('tracking_type', 'TIME_ONLY'),
                        'custom_categories': item.get('custom_categories', []),
                        'html_layout': item.get('html_layout', ''),
                        'ai_prompt_schema': item.get('ai_prompt_schema'),
                        'custom_extraction_prompt': item.get('custom_extraction_prompt'),
                        'l1_signature_role': item.get('l1_signature_role'),
                        'l2_signature_role': item.get('l2_signature_role'),
                    }
                )
                
                # --- THE V1 FIX FOR STATIC DOCS ---
                # Only spawn the anchor if this template was just created (not just updated)
                if created and doc_type == 'PRINT_ONLY':
                    report = DailyReport.objects.create(
                        template=template,
                        report_date=date.today(),
                        latest_data={},
                        status='DRAFT'
                    )
                    
                    ReportVersion.objects.create(
                        daily_report=report,
                        version_number=1,
                        action_type='MANUAL_EDIT',
                        # CHANGED: Removed the word "Imported" to make it look native
                        update_reason="Initial Static Document Deployed", 
                        data_snapshot={},
                        modified_by=request.user
                    )
                
                imported_count += 1
                
            messages.success(request, f"Success! {imported_count} templates synced to Master Database.")
        except Exception as e:
            messages.error(request, f"Sync Failed: Ensure the file is a valid JSON export. ({str(e)})")
            
    return redirect('records:template_sync')