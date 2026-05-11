import json
import csv
import base64
import requests
import os
import zipfile
import calendar
import tempfile
import random
from django.core.mail import send_mail
from django.contrib.auth import update_session_auth_hash
from io import BytesIO, StringIO
from datetime import date, timedelta, datetime
from django.template.loader import render_to_string
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db import transaction
from django.db.models import Sum, Q, Count, F
from django.views.decorators.http import require_POST
from django.conf import settings
from django.core.management import call_command
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl.utils import get_column_letter

# === HUB SECURITY IMPORTS ===
from hub.decorators import require_app_access, require_feature_access

from .models import Center, StockSheet, MasterItem, ItemAlias, StockEntry, CenterMasterItem, CenterItemAlias, WhatsAppReceipt, PrintTemplate, PrintTemplateItem, CenterClassification, TargetReportConfig, TargetReportColumn
from .utils import analyze_with_gemini

# ==========================================
# 1. DASHBOARD & UPLOAD
# ==========================================

@require_app_access('inventory')
@ensure_csrf_cookie 
def dashboard(request):
    centers = Center.objects.all().order_by('name')
    return render(request, 'inventory/index.html', {'centers': centers})

@require_app_access('inventory')
def api_check_uploads(request):
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'ids': []})
    
    completed_ids = StockSheet.objects.filter(date=date_str).values_list('center_id', flat=True)
    return JsonResponse({'ids': list(completed_ids)})

@require_app_access('inventory')
@require_POST
def api_attach_previous_report(request):
    """Duplicates the most recent previous report for a center to the target date."""
    try:
        data = json.loads(request.body)
        center_id = data.get('center_id')
        target_date = data.get('date')

        if not center_id or not target_date:
            return JsonResponse({'status': 'error', 'message': 'Missing center or date.'})

        target_date_obj = datetime.strptime(target_date, '%Y-%m-%d').date()

        prev_sheet = StockSheet.objects.filter(center_id=center_id, date__lt=target_date_obj).order_by('-date').first()
        
        if not prev_sheet:
            return JsonResponse({'status': 'error', 'message': 'No previous report found to carry over.'})

        with transaction.atomic():
            sheet, created = StockSheet.objects.get_or_create(center_id=center_id, date=target_date_obj)
            
            sheet.image = prev_sheet.image
            sheet.is_carried_over = True
            sheet.copied_from_date = prev_sheet.date
            sheet.save()

            sheet.entries.all().delete()
            
            new_entries = []
            for entry in prev_sheet.entries.all():
                new_entries.append(StockEntry(
                    sheet=sheet,
                    master_item=entry.master_item,
                    sheet_category=entry.sheet_category,
                    closing_balance=entry.closing_balance,
                    raw_name_scanned=entry.raw_name_scanned
                ))
            StockEntry.objects.bulk_create(new_entries)

        return JsonResponse({'status': 'success', 'message': f"Copied report from {prev_sheet.date.strftime('%d %b')}"})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_app_access('inventory')
@require_POST
def save_stock(request):
    try:
        center_id = request.POST.get('center_id')
        date_str = request.POST.get('date')
        items_str = request.POST.get('items', '[]')
        image = request.FILES.get('stock_image')

        if not center_id or not date_str:
            return JsonResponse({'status': 'error', 'message': 'Missing center or date'})

        center = get_object_or_404(Center, id=center_id)
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        items = json.loads(items_str)

        with transaction.atomic():
            sheet, created = StockSheet.objects.get_or_create(
                center=center, 
                date=date_obj,
                defaults={'image': image}
            )
            if not created and image:
                sheet.image = image
                sheet.save()

            sheet.entries.all().delete()

            existing_center_masters = {
                cm.name: cm for cm in CenterMasterItem.objects.filter(center=center)
            }

            for item in items:
                raw_name = item.get('material', '').strip().upper()
                original_scanned = item.get('original_scanned', raw_name).strip().upper()
                raw_cat = item.get('category', '').strip() 
                
                opening = item.get('opening_balance', 0)
                inward = item.get('inward', 0)
                production = item.get('production', 0)
                dispatch = item.get('dispatch', 0)
                closing = item.get('closing_balance', 0)
                
                save_alias = item.get('save_alias', False)
                
                if not raw_name: continue

                global_master, _ = MasterItem.objects.get_or_create(name=raw_name, defaults={'category': raw_cat})

                center_master = existing_center_masters.get(raw_name)
                if not center_master:
                    center_master = CenterMasterItem.objects.create(
                        center=center, name=raw_name, category=raw_cat, display_order=len(existing_center_masters)
                    )
                    existing_center_masters[raw_name] = center_master

                if save_alias and original_scanned and original_scanned != raw_name:
                    CenterItemAlias.objects.get_or_create(
                        center=center, scanned_name=original_scanned, defaults={'mapped_to': center_master}
                    )

                StockEntry.objects.create(
                    sheet=sheet, master_item=global_master, sheet_category=raw_cat,
                    opening_balance=opening, inward=inward, production=production, 
                    dispatch=dispatch, closing_balance=closing,
                    raw_name_scanned=original_scanned
                )

        return JsonResponse({'status': 'success'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_app_access('inventory')
def api_get_previous_closing(request):
    """Fetches the final closing balances from the EXACT previous date (T-1)."""
    center_id = request.GET.get('center_id')
    date_str = request.GET.get('date')
    
    if not center_id or not date_str: 
        return JsonResponse({'status': 'error', 'balances': {}})
        
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    exact_prev_date = date_obj - timedelta(days=1)
    prev_sheet = StockSheet.objects.filter(center_id=center_id, date=exact_prev_date).first()
    
    if not prev_sheet: 
        return JsonResponse({
            'status': 'success', 
            'balances': {}, 
            'exact_match': False,
            'message': f"No report found for yesterday ({exact_prev_date.strftime('%d-%b-%Y')}). All carried forward balances are 0."
        })
        
    balances = {}
    for entry in prev_sheet.entries.all():
        balances[entry.master_item.name.upper()] = float(entry.closing_balance)
        
    return JsonResponse({
        'status': 'success', 
        'balances': balances, 
        'exact_match': True
    })

# ==========================================
# 2. MASTER ITEMS
# ==========================================

@require_app_access('inventory')
def master_list(request):
    items = MasterItem.objects.prefetch_related('aliases').all().order_by('name')
    return render(request, 'inventory/master_list.html', {'items': items})

# ==========================================
# 3. CENTERS MANAGEMENT
# ==========================================

@require_app_access('inventory')
def centers_list(request):
    centers_data = []
    all_centers = Center.objects.all().order_by('name')

    for center in all_centers:
        latest_sheet = StockSheet.objects.filter(center=center).order_by('-date').first()
        
        if latest_sheet:
            raw_total = latest_sheet.entries.filter(sheet_category__icontains='Raw').aggregate(t=Sum('closing_balance'))['t'] or 0
            finished_total = latest_sheet.entries.filter(sheet_category__icontains='Finish').aggregate(t=Sum('closing_balance'))['t'] or 0
            last_date = latest_sheet.date
            sheet_id = latest_sheet.id
        else:
            raw_total = 0
            finished_total = 0
            last_date = None
            sheet_id = None

        centers_data.append({
            'obj': center,
            'raw_total': int(raw_total),
            'finished_total': int(finished_total),
            'last_date': last_date,
            'sheet_id': sheet_id
        })

    return render(request, 'inventory/centers_list.html', {'centers': centers_data})

@require_app_access('inventory')
def manage_centers(request):
    # FIXED: Using the exact model names from your original code
    centers = Center.objects.select_related('classification').all().order_by('name')
    classifications = CenterClassification.objects.all().order_by('name')
    
    # Granular Security Evaluation
    can_manage_class = False
    can_delete_center = False
    
    if request.user.is_superuser:
        can_manage_class = True
        can_delete_center = True
    elif hasattr(request.user, 'nexus_profile'):
        roles = request.user.nexus_profile.roles.all()
        can_manage_class = roles.filter(allowed_features__code_name='can_manage_classifications').exists()
        can_delete_center = roles.filter(allowed_features__code_name='can_delete_center').exists()
        
    return render(request, 'inventory/manage_centers.html', {
        'centers': centers,
        'classifications': classifications,
        'can_manage_class': can_manage_class,
        'can_delete_center': can_delete_center
    })

@require_app_access('inventory')
@require_POST
def update_center_api(request):
    try:
        data = json.loads(request.body)
        action = data.get('action')

        # 1. DELETE ACTION (Restricted)
        if action == 'delete':
            if not request.user.is_superuser:
                has_delete = request.user.nexus_profile.roles.filter(allowed_features__code_name='can_delete_center').exists()
                if not has_delete:
                    return JsonResponse({'status': 'error', 'message': 'You lack permissions to delete centers.'}, status=403)
            
            center_id = data.get('id')
            Center.objects.filter(id=center_id).delete()
            return JsonResponse({'status': 'success'})

        # 2. ADD ACTION (Open to all staff)
        if action == 'add':
            name = data.get('name', '').strip()
            state = data.get('state', '').strip()
            category = data.get('category', '').strip() 
            classification_id = data.get('classification_id') 

            if not name or not state:
                return JsonResponse({'status': 'error', 'message': 'Center Name and State are mandatory!'})

            if category not in ['HD', 'LD']: category = None
            cls_obj = CenterClassification.objects.filter(id=classification_id).first() if classification_id else None

            if Center.objects.filter(name__iexact=name).exists():
                 return JsonResponse({'status': 'error', 'message': 'Center with this name already exists'})
            
            Center.objects.create(name=name, state=state, category=category, classification=cls_obj) 
            return JsonResponse({'status': 'success'})
            
        # 3. BULK UPDATE ACTION (Open to all staff)
        if action == 'bulk_update':
            updates = data.get('updates', [])
            for item in updates:
                c_id = item.get('id')
                name = item.get('name', '').strip()
                state = item.get('state', '').strip()
                category = item.get('category', '')
                class_id = item.get('classification_id')
                
                if category not in ['HD', 'LD']: category = None
                cls_obj = CenterClassification.objects.filter(id=class_id).first() if class_id else None
                
                # Update row directly
                Center.objects.filter(id=c_id).update(
                    name=name, state=state, category=category, classification=cls_obj
                )
                
            return JsonResponse({'status': 'success'})

        return JsonResponse({'status': 'error', 'message': 'Invalid action code.'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@require_app_access('inventory')
@require_POST
def manage_classification_api(request):
    # SECURITY GATE: Check permission inside the function instead of the decorator
    if not request.user.is_superuser:
        if not request.user.nexus_profile.roles.filter(allowed_features__code_name='can_manage_classifications').exists():
            return JsonResponse({'status': 'error', 'message': 'You lack permissions to manage classifications.'}, status=403)

    try:
        data = json.loads(request.body)
        action = data.get('action')
        
        if action == 'add':
            name = data.get('name', '').strip()
            if not name: return JsonResponse({'status': 'error', 'message': 'Name required'})
            if CenterClassification.objects.filter(name__iexact=name).exists():
                return JsonResponse({'status': 'error', 'message': 'Classification already exists'})
            CenterClassification.objects.create(name=name)
            
        elif action == 'delete':
            c_id = data.get('id')
            CenterClassification.objects.filter(id=c_id).delete()
            
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
        
# ==========================================
# 4. CENTER HISTORY
# ==========================================

@require_app_access('inventory')
def center_history(request, center_id):
    center = get_object_or_404(Center, id=center_id)
    
    sort_param = request.GET.get('sort', 'desc') 
    if sort_param == 'asc':
        sheets = StockSheet.objects.filter(center=center).order_by('date')
        next_sort = 'desc'
        sort_icon = 'fa-arrow-up-1-9'
    else:
        sheets = StockSheet.objects.filter(center=center).order_by('-date')
        next_sort = 'asc'
        sort_icon = 'fa-arrow-down-9-1'

    history_data = []
    for sheet in sheets:
        raw = sheet.entries.filter(sheet_category__icontains='Raw').aggregate(t=Sum('closing_balance'))['t'] or 0
        finished = sheet.entries.filter(sheet_category__icontains='Finish').aggregate(t=Sum('closing_balance'))['t'] or 0
        
        history_data.append({
            'sheet': sheet,
            'raw_total': int(raw),
            'finished_total': int(finished)
        })

    return render(request, 'inventory/center_history.html', {
        'center': center,
        'history': history_data,
        'next_sort': next_sort,
        'sort_icon': sort_icon
    })


# ==========================================
# 5. REPORT DETAIL & SINGLE EXCEL EXPORT
# ==========================================

@require_app_access('inventory')
def report_detail(request, sheet_id):
    sheet = get_object_or_404(StockSheet, id=sheet_id)
    entries = sheet.entries.select_related('master_item').all().order_by('sheet_category', 'master_item__name')
    
    raw_total = entries.filter(sheet_category__icontains='Raw').aggregate(s=Sum('closing_balance'))['s'] or 0
    finished_total = entries.filter(sheet_category__icontains='Finish').aggregate(s=Sum('closing_balance'))['s'] or 0
    grand_total = raw_total + finished_total

    return render(request, 'inventory/report_detail.html', {
        'sheet': sheet, 
        'entries': entries,
        'raw_total': int(raw_total), 
        'finished_total': int(finished_total),
        'grand_total': int(grand_total)
    })

@require_app_access('inventory')
def update_report(request, sheet_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'})
        
    try:
        sheet = get_object_or_404(StockSheet, id=sheet_id)
        
        # Check if they are moving the report to a new date
        new_date = request.POST.get('date')
        if new_date and new_date != str(sheet.date):
            if StockSheet.objects.filter(center=sheet.center, date=new_date).exclude(id=sheet.id).exists():
                return JsonResponse({'status': 'error', 'message': f'A report for {new_date} already exists for this center!'})
            sheet.date = new_date
            
        sheet.is_carried_over = False
        sheet.copied_from_date = None
        
        if request.FILES.get('stock_image'):
            sheet.image = request.FILES['stock_image']
            
        sheet.save()

        items = json.loads(request.POST.get('items', '[]'))
        
        with transaction.atomic():
            sheet.entries.all().delete()
            
            for item in items:
                raw_name = item.get('material', '').strip()
                raw_cat = item.get('category', '').strip()
                
                # EXTRACT ALL 4 COLUMNS
                op = float(item.get('opening_balance', 0))
                inw = float(item.get('inward', 0))
                disp = float(item.get('dispatch', 0))
                clo = float(item.get('closing_balance', 0))
                
                if not raw_name: continue
                clean_name = " ".join(raw_name.split()).lower()

                alias_obj = ItemAlias.objects.filter(alias_name=clean_name).first()
                if alias_obj:
                    master_item = alias_obj.master_item
                else:
                    master_item = MasterItem.objects.filter(name=clean_name).first()
                    if not master_item:
                        master_item = MasterItem.objects.create(name=clean_name, category=raw_cat)

                # SAVE ALL 4 COLUMNS TO THE DATABASE
                StockEntry.objects.create(
                    sheet=sheet,
                    master_item=master_item,
                    sheet_category=raw_cat,
                    opening_balance=op,
                    inward=inw,
                    dispatch=disp,
                    closing_balance=clo,
                    raw_name_scanned=raw_name
                )
                
        return JsonResponse({'status': 'success', 'message': 'Report updated!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@require_app_access('inventory')
def api_navigate_report(request, center_id):
    """Engine for the Prev/Next/Jump time-travel buttons."""
    date_str = request.GET.get('date')
    action = request.GET.get('action') # 'exact', 'prev', 'next'
    
    current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    if action == 'exact':
        sheet = StockSheet.objects.filter(center_id=center_id, date=current_date).first()
    elif action == 'prev':
        sheet = StockSheet.objects.filter(center_id=center_id, date__lt=current_date).order_by('-date').first()
    elif action == 'next':
        sheet = StockSheet.objects.filter(center_id=center_id, date__gt=current_date).order_by('date').first()
        
    if sheet:
        return JsonResponse({'status': 'success', 'sheet_id': sheet.id})
    return JsonResponse({'status': 'error', 'message': 'No report found.'})
    
@require_app_access('inventory')
def download_sheet_csv(request, sheet_id):
    sheet = get_object_or_404(StockSheet, id=sheet_id)
    entries = sheet.entries.all().select_related('master_item').order_by('sheet_category', 'master_item__name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Stock_{sheet.date.strftime('%d-%b')}"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="64748B")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0')
    )

    ws['A1'] = f"{sheet.center.name.upper()} - STOCK REPORT"
    ws['A1'].font = title_font
    
    ws['A2'] = f"Recorded Date: {sheet.date.strftime('%d %B %Y')}"
    ws['A2'].font = subtitle_font

    headers = ["Type", "Material Name", "Closing Balance"]
    ws.append([]) 
    ws.append(headers) 

    for col_num in range(1, 4):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align if col_num != 2 else left_align
        cell.border = thin_border

    row_num = 5
    for entry in entries:
        if entry.sheet_category == "Raw Material": type_badge = "RM"
        elif entry.sheet_category == "Finished Goods": type_badge = "FG"
        else: type_badge = entry.sheet_category

        ws.cell(row=row_num, column=1, value=type_badge).alignment = center_align
        ws.cell(row=row_num, column=2, value=entry.master_item.name).alignment = left_align
        ws.cell(row=row_num, column=3, value=entry.closing_balance).alignment = right_align
        
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).border = thin_border
            
        row_num += 1

    ws.auto_filter.ref = f"A4:C{row_num-1}"
    ws.freeze_panes = ws['A5']

    ws.column_dimensions['A'].width = 12 
    ws.column_dimensions['B'].width = 45  
    ws.column_dimensions['C'].width = 20 

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{sheet.center.name.replace(' ', '_')}_Stock_{sheet.date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# ==========================================
# 6. DAILY MATRIX REPORTS
# ==========================================

def get_daily_matrix(date_str):
    all_centers = Center.objects.all().order_by('name')
    
    center_headers = []
    for center in all_centers:
        sheet = StockSheet.objects.filter(center=center, date=date_str).first()
        center_headers.append({
            'obj': center,
            'sheet_id': sheet.id if sheet else None,
            'image_url': f"../media/{sheet.image.name}" if sheet and sheet.image else None,
            'is_carried_over': sheet.is_carried_over if sheet else False,
            'copied_from_date': sheet.copied_from_date.strftime('%d %b %Y') if sheet and sheet.copied_from_date else None
        })

    center_summaries = []
    grand_rm = 0
    grand_fg = 0
    
    for data in center_headers:
        sheet_id = data['sheet_id']
        
        if sheet_id:
            rm = StockEntry.objects.filter(sheet_id=sheet_id, sheet_category__icontains='Raw').aggregate(s=Sum('closing_balance'))['s'] or 0
            fg = StockEntry.objects.filter(sheet_id=sheet_id, sheet_category__icontains='Finish').aggregate(s=Sum('closing_balance'))['s'] or 0
        else:
            rm = 0
            fg = 0
            
        grand_rm += rm
        grand_fg += fg
        
        center_summaries.append({
            'rm': int(rm),
            'fg': int(fg),
            'total': int(rm + fg)
        })

    entries = StockEntry.objects.filter(
        sheet__date=date_str
    ).select_related('sheet', 'master_item')
    
    stock_map = {}
    unique_item_cats = set() 
    
    for e in entries:
        cat = 'Raw Material' if 'Raw' in e.sheet_category else 'Finished Goods'
        key = f"{e.sheet.center.id}_{e.master_item.id}_{cat}"
        
        if key not in stock_map:
            stock_map[key] = {'op': 0, 'inw': 0, 'disp': 0, 'clo': 0}
            
        # AGGREGATE ALL 4 COLUMNS
        stock_map[key]['op'] += e.opening_balance
        stock_map[key]['inw'] += e.inward
        stock_map[key]['disp'] += e.dispatch
        stock_map[key]['clo'] += e.closing_balance
        
        unique_item_cats.add((e.master_item, cat))

    sorted_item_cats = sorted(list(unique_item_cats), key=lambda x: (x[1], x[0].name.lower()))

    detailed_rows = []
    for item, cat in sorted_item_cats:
        row_cols = []
        
        for data in center_headers:
            center = data['obj']
            key = f"{center.id}_{item.id}_{cat}"
            val_dict = stock_map.get(key) 
            
            if val_dict:
                row_cols.append({
                    'op': int(val_dict['op']),
                    'inw': int(val_dict['inw']),
                    'disp': int(val_dict['disp']),
                    'clo': int(val_dict['clo'])
                })
            else:
                row_cols.append(None)
                
        detailed_rows.append({
            'item': item,
            'category': cat,
            'cols': row_cols
        })

    return {
        'centers': center_headers, 
        'summaries': center_summaries,
        'grand_rm': int(grand_rm),
        'grand_fg': int(grand_fg),
        'detailed_rows': detailed_rows
    }

@require_app_access('inventory')
def daily_reports_list(request):
    dates = StockSheet.objects.dates('date', 'day', order='DESC')
    return render(request, 'inventory/daily_reports_list.html', {'dates': dates})

@require_app_access('inventory')
def daily_report_detail(request, date_str):
    data = get_daily_matrix(date_str)
    classifications = CenterClassification.objects.all().order_by('name')
    return render(request, 'inventory/daily_report_detail.html', {
        'date': date_str,
        'classifications': classifications,
        **data
    })

# ==========================================
# 7. DAILY MATRIX EXCEL EXPORT (FIXED LAYOUT)
# ==========================================

@require_app_access('inventory')
def download_daily_report(request, date_str):
    data = get_daily_matrix(date_str)
    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    wb = openpyxl.Workbook()
    
    # --- 1. DETERMINE EXPORT TYPE ---
    export_type = request.POST.get('export_type', 'closing')
    type_map = {
        'opening': 'op',
        'inward': 'inw',
        'production': 'inw',
        'dispatch': 'disp',
        'closing': 'clo'
    }
    val_key = type_map.get(export_type, 'clo')

    views = []

    if request.method == 'POST':
        # Collection Centers & Subsets
        if request.POST.get('sheet_collection'):
            views.append({
                "title": "Collection Centers", 
                "filter_func": lambda c: c.classification and 'collection' in c.classification.name.lower()
            })
        if request.POST.get('sheet_ld_collection'):
            views.append({
                "title": "LD Collection", 
                "filter_func": lambda c: c.classification and 'collection' in c.classification.name.lower() and c.category == 'LD'
            })
        if request.POST.get('sheet_hd_collection'):
            views.append({
                "title": "HD Collection", 
                "filter_func": lambda c: c.classification and 'collection' in c.classification.name.lower() and c.category == 'HD'
            })
            
        # Dynamic Operational Groups
        classifications = CenterClassification.objects.exclude(name__icontains='collection').order_by('name')
        for cls in classifications:
            if request.POST.get(f'sheet_class_{cls.id}'):
                def make_filter(cls_id):
                    return lambda c: c.classification and c.classification.id == cls_id
                views.append({
                    "title": cls.name[:31], 
                    "filter_func": make_filter(cls.id)
                })
    else:
        views.append({"title": "All Centers", "filter_func": lambda c: True})

    if not views:
        views.append({"title": "All Centers", "filter_func": lambda c: True})
        
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(bold=True, name='Calibri', size=11)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    border_thin = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))
    
    default_sheet = wb.active
    first_sheet = True
    
    for view in views:
        if first_sheet:
            ws = default_sheet
            ws.title = view["title"]
            first_sheet = False
        else:
            ws = wb.create_sheet(title=view["title"])
            
        # Map original center indices to keep alignment safe
        filtered_centers_indices = []
        filtered_centers = []
        for idx, c in enumerate(data['centers']):
            if view["filter_func"](c['obj']):
                filtered_centers_indices.append(idx)
                filtered_centers.append(c)
                
        if not filtered_centers:
            ws.cell(row=1, column=1, value=f"No {view['title']} found for this date.").font = font_bold
            continue

        # --- 2. DYNAMIC ROW FILTERING ---
        filtered_rows = []
        for row in data['detailed_rows']:
            # Inward = RM Only | Production = FG Only
            if export_type == 'inward' and row['category'] != 'Raw Material':
                continue
            if export_type == 'production' and row['category'] != 'Finished Goods':
                continue
            filtered_rows.append(row)

        # --- 3. DYNAMIC COLUMN SUMMARIES (Calculated on the fly based on export_type) ---
        dynamic_summaries = []
        for original_idx in filtered_centers_indices:
            c_rm, c_fg = 0, 0
            for row in filtered_rows:
                val_dict = row['cols'][original_idx]
                if val_dict:
                    val = val_dict.get(val_key, 0)
                    if row['category'] == 'Raw Material':
                        c_rm += val
                    else:
                        c_fg += val
            dynamic_summaries.append({'rm': c_rm, 'fg': c_fg, 'total': c_rm + c_fg})

        view_grand_rm = sum(s['rm'] for s in dynamic_summaries)
        view_grand_fg = sum(s['fg'] for s in dynamic_summaries)
        view_grand_total = view_grand_rm + view_grand_fg

        # Build Headers
        title_prefix = "INWARD" if export_type == 'inward' else export_type.upper()
        ws.cell(row=1, column=1, value=f"{view['title'].upper()} - {title_prefix} REPORT ({date_str})").font = Font(bold=True, size=16, color="1E293B")
        
        center_names = []
        for c in filtered_centers:
            header_text = c['obj'].name
            if c.get('is_carried_over') and c.get('copied_from_date'):
                header_text += f"\n(From {c['copied_from_date']})"
            center_names.append(header_text)
            
        headers = ['Sr No', 'Type', 'Material Name', 'Total'] + center_names
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center if col_num != 3 else align_left
            cell.border = border_thin

        # RM Summary Row
        ws.cell(row=4, column=1, value="").border = border_thin
        ws.cell(row=4, column=2, value="RM").alignment = align_center
        ws.cell(row=4, column=2).border = border_thin
        ws.cell(row=4, column=3, value="RAW MATERIAL TOTAL").font = font_bold
        ws.cell(row=4, column=3).border = border_thin
        ws.cell(row=4, column=4, value=view_grand_rm).font = font_bold
        ws.cell(row=4, column=4).alignment = align_center
        ws.cell(row=4, column=4).border = border_thin
        for i, s in enumerate(dynamic_summaries):
            val = s['rm'] if s['rm'] else "-"
            cell = ws.cell(row=4, column=i+5, value=val)
            cell.alignment = align_center
            cell.border = border_thin

        # FG Summary Row
        ws.cell(row=5, column=1, value="").border = border_thin
        ws.cell(row=5, column=2, value="FG").alignment = align_center
        ws.cell(row=5, column=2).border = border_thin
        ws.cell(row=5, column=3, value="FINISHED GOODS TOTAL").font = font_bold
        ws.cell(row=5, column=3).border = border_thin
        ws.cell(row=5, column=4, value=view_grand_fg).font = font_bold
        ws.cell(row=5, column=4).alignment = align_center
        ws.cell(row=5, column=4).border = border_thin
        for i, s in enumerate(dynamic_summaries):
            val = s['fg'] if s['fg'] else "-"
            cell = ws.cell(row=5, column=i+5, value=val)
            cell.alignment = align_center
            cell.border = border_thin

        # Grand Total Row
        ws.cell(row=6, column=1, value="").border = border_thin
        ws.cell(row=6, column=2, value="ALL").alignment = align_center
        ws.cell(row=6, column=2).border = border_thin
        ws.cell(row=6, column=3, value="GRAND TOTAL").font = Font(bold=True, color="4F46E5")
        ws.cell(row=6, column=3).border = border_thin
        ws.cell(row=6, column=4, value=view_grand_total).font = Font(bold=True, color="4F46E5")
        ws.cell(row=6, column=4).alignment = align_center
        ws.cell(row=6, column=4).border = border_thin
        for i, s in enumerate(dynamic_summaries):
            val = s['total'] if s['total'] else "-"
            cell = ws.cell(row=6, column=i+5, value=val)
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border_thin

        # --- 4. POPULATE DETAILED ROWS ---
        current_row = 7
        sr_no = 1
        
        for row_data in filtered_rows:
            row_total = 0
            filtered_cols = []
            
            for original_idx in filtered_centers_indices:
                val_dict = row_data['cols'][original_idx]
                val = val_dict.get(val_key, 0) if val_dict else 0
                filtered_cols.append(val)
                row_total += val
                
            if row_total == 0: continue # Skip empty rows for this specific report

            type_badge = "RM" if row_data['category'] == 'Raw Material' else "FG"

            ws.cell(row=current_row, column=1, value=sr_no).alignment = align_center
            ws.cell(row=current_row, column=2, value=type_badge).alignment = align_center
            ws.cell(row=current_row, column=3, value=row_data['item'].name.upper()).alignment = align_left
            
            cell_total = ws.cell(row=current_row, column=4, value=row_total)
            cell_total.font = font_bold
            cell_total.alignment = align_center

            for i, val in enumerate(filtered_cols):
                display_val = val if val else "-"
                cell = ws.cell(row=current_row, column=i+5, value=display_val)
                cell.alignment = align_center
                cell.border = border_thin
                
            for c_idx in range(1, 5):
                ws.cell(row=current_row, column=c_idx).border = border_thin

            current_row += 1
            sr_no += 1

        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A3:{last_col_letter}{current_row-1}"
        ws.freeze_panes = 'D7'

        ws.column_dimensions['A'].width = 8   
        ws.column_dimensions['B'].width = 8   
        ws.column_dimensions['C'].width = 35  
        ws.column_dimensions['D'].width = 12  
        for i in range(5, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14

    # Dynamically name the file based on the report type!
    file_prefix = export_type.capitalize()
    filename = f"{file_prefix}_Report_{date_obj.strftime('%d_%b_%Y')}.xlsx"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ==========================================
# 9. API PROXIES
# ==========================================

@require_app_access('inventory')
@require_POST
def api_analyze_sheet(request):
    try:
        data = json.loads(request.body)
        mime = data.get('mime_type')
        b64_data = data.get('data')
        
        if not mime or not b64_data:
            return JsonResponse({'error': 'Missing file data'}, status=400)
            
        result = analyze_with_gemini(mime, b64_data)
        
        if 'error' in result:
             return JsonResponse(result, status=500)
             
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
# ==========================================
# 10. NEXUS PUSH & HTML SYNC (SUPERUSER ONLY)
# ==========================================

@require_app_access('inventory')
def download_daily_html(request, date_str):
    data = get_daily_matrix(date_str)
    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    
    prev_d = StockSheet.objects.filter(date__lt=date_obj).order_by('-date').values_list('date', flat=True).first()
    next_d = StockSheet.objects.filter(date__gt=date_obj).order_by('date').values_list('date', flat=True).first()
    
    prev_url = f"stockreport-{prev_d.strftime('%d%b').lower()}.html" if prev_d else None
    next_url = f"stockreport-{next_d.strftime('%d%b').lower()}.html" if next_d else None
    
    html_content = render_to_string('inventory/daily_report_standalone.html', {
        'date': date_obj, 'prev_url': prev_url, 'next_url': next_url, **data
    })
    
    filename = f"stockreport-{date_obj.strftime('%d%b').lower()}.html"
    response = HttpResponse(html_content, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@require_app_access('inventory')
def api_get_nexus_manifest(request):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Superuser access required.'}, status=403)

    dates = list(StockSheet.objects.values_list('date', flat=True).distinct().order_by('date'))
    date_strs = [d.strftime('%Y-%m-%d') for d in dates]

    media_files = []
    media_root = settings.MEDIA_ROOT
    if os.path.exists(media_root):
        for root, dirs, files in os.walk(media_root):
            for file in files:
                file_path = os.path.join(root, file)
                rel_path = os.path.relpath(file_path, media_root).replace('\\', '/')
                media_files.append(rel_path)

    return JsonResponse({'dates': date_strs, 'media': media_files})

@require_app_access('inventory')
@require_POST
def api_push_nexus_item(request):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Superuser access required.'}, status=403)

    try:
        data = json.loads(request.body)
        item_type = data.get('type', 'report') 
        item_val = data.get('value')
        
        # SECURED: Pull from environment variables, never hardcode tokens!
        GITHUB_TOKEN = os.environ.get('GITHUB_NEXUS_TOKEN', '')
        if not GITHUB_TOKEN:
            return JsonResponse({'status': 'error', 'message': 'GitHub Token not configured on server.'}, status=500)
            
        REPO_OWNER = "kkaranmistry"   
        REPO_NAME = "nexus-view"    

        headers = {
            "Authorization": f"Bearer {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json"
        }

        if item_type == 'report' or not data.get('type'):
            date_str = item_val
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            matrix_data = get_daily_matrix(date_str)
            
            prev_d = StockSheet.objects.filter(date__lt=date_obj).order_by('-date').values_list('date', flat=True).first()
            next_d = StockSheet.objects.filter(date__gt=date_obj).order_by('date').values_list('date', flat=True).first()
            
            prev_url = f"stockreport-{prev_d.strftime('%d%b').lower()}.html" if prev_d else None
            next_url = f"stockreport-{next_d.strftime('%d%b').lower()}.html" if next_d else None
            
            html_content = render_to_string('inventory/daily_report_standalone.html', {
                'date': date_obj, 'prev_url': prev_url, 'next_url': next_url, **matrix_data
            })
            
            encoded_content = base64.b64encode(html_content.encode('utf-8')).decode('utf-8')
            file_path = f"reports/stockreport-{date_obj.strftime('%d%b').lower()}.html"
            
        elif item_type == 'media':
            abs_path = os.path.join(settings.MEDIA_ROOT, item_val)
            with open(abs_path, 'rb') as f:
                encoded_content = base64.b64encode(f.read()).decode('utf-8')
            file_path = f"media/{item_val}"
            
        url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{file_path}"
        
        get_resp = requests.get(url, headers=headers)
        payload = {"message": f"[skip ci] Auto-sync: {file_path}", "content": encoded_content}
        if get_resp.status_code == 200:
            payload["sha"] = get_resp.json().get('sha')
            
        put_resp = requests.put(url, headers=headers, json=payload)
        
        if put_resp.status_code in [200, 201]:
            return JsonResponse({"status": "success"})
        else:
            return JsonResponse({"status": "error", "message": put_resp.json().get('message')})

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@require_app_access('inventory')
@require_POST
def api_push_to_nexus(request, date_str):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Superuser access required.'}, status=403)

    try:
        request._body = json.dumps({'type': 'report', 'value': date_str}).encode('utf-8')
        return api_push_nexus_item(request)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})


# ==========================================
# 11. WHATSAPP TRACKER
# ==========================================

@require_app_access('inventory')
def api_whatsapp_status(request):
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'received_ids': []})
        
    receipts = WhatsAppReceipt.objects.filter(date=date_str).values_list('center_id', flat=True)
    return JsonResponse({'received_ids': list(receipts)})

@require_app_access('inventory')
@require_POST
def api_whatsapp_toggle(request):
    try:
        data = json.loads(request.body)
        center_id = data.get('center_id')
        date_str = data.get('date')
        
        center = get_object_or_404(Center, id=center_id)
        
        receipt, created = WhatsAppReceipt.objects.get_or_create(center=center, date=date_str)
        
        if not created:
            receipt.delete()
            is_received = False
        else:
            is_received = True
            
        return JsonResponse({'status': 'success', 'is_received': is_received})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
@require_app_access('inventory')
def whatsapp_tracker(request):
    centers = Center.objects.all().order_by('name')
    return render(request, 'inventory/whatsapp_tracker.html', {'centers': centers})


# ==========================================
# 12. USER PROFILE & SECURITY
# ==========================================

@require_app_access('inventory')
def user_profile(request):
    # Determine if they get the Hub button (Superuser OR > 1 app)
    has_multiple_apps = False
    if request.user.is_superuser:
        has_multiple_apps = True
    elif hasattr(request.user, 'nexus_profile'):
        has_multiple_apps = request.user.nexus_profile.allowed_apps.count() > 1
        
    return render(request, 'inventory/profile.html', {
        'has_multiple_apps': has_multiple_apps
    })

@require_app_access('inventory')
@require_POST
def api_update_profile(request):
    try:
        data = json.loads(request.body)
        request.user.first_name = data.get('first_name', '').strip()
        request.user.last_name = data.get('last_name', '').strip()
        request.user.email = data.get('email', '').strip()
        request.user.save()
        
        return JsonResponse({'status': 'success', 'message': 'Profile updated successfully!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_app_access('inventory')
@require_POST
def api_request_otp(request):
    if not request.user.email:
        return JsonResponse({'status': 'error', 'message': 'No email address saved to this account.'})
    
    # Force a strict 6-digit string (e.g., "048291")
    import random
    otp = f"{random.randint(0, 999999):06d}"
    request.session['password_reset_otp'] = otp
    
    try:
        from django.core.mail import send_mail
        from django.conf import settings
        
        send_mail(
            subject='Nexus Workspace - Security Verification Code',
            message=f'Your 6-digit verification code is: {otp}\n\nIf you did not request this change, please ignore this email or notify your administrator.',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[request.user.email],
            fail_silently=False,
        )
        return JsonResponse({'status': 'success'})
    except Exception as e:
        print(f"EMAIL ERROR: {str(e)}") 
        return JsonResponse({'status': 'error', 'message': 'Failed to send email. Check console.'})
        
@require_app_access('inventory')
@require_POST
def api_verify_otp(request):
    try:
        data = json.loads(request.body)
        submitted_otp = data.get('otp', '').strip()
        new_password = data.get('new_password', '')
        
        stored_otp = request.session.get('password_reset_otp')
        
        if not stored_otp or submitted_otp != stored_otp:
            return JsonResponse({'status': 'error', 'message': 'Invalid or expired OTP.'})
            
        if len(new_password) < 8:
            return JsonResponse({'status': 'error', 'message': 'Password must be at least 8 characters.'})
            
        # 1. Change the password
        request.user.set_password(new_password)
        request.user.save()
        
        # 2. Keep the user logged in! (Changing password usually forces a logout in Django)
        update_session_auth_hash(request, request.user)
        
        # 3. Destroy the OTP so it can't be reused
        del request.session['password_reset_otp']
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
        
# ==========================================
# 13. MOVEMENT & PRODUCTION ENGINE
# ==========================================

def get_report_data(start_date, end_date, report_type):
    results = []
    
    # Base queryset filtering by date range
    entries = StockEntry.objects.filter(
        sheet__date__range=[start_date, end_date]
    ).select_related('sheet__center', 'master_item').order_by('sheet__date', 'sheet__center__name')

    # Apply specific filters based on report type
    if report_type == 'inward':
        entries = entries.filter(sheet_category__icontains='Raw', inward__gt=0)
    elif report_type == 'production':
        entries = entries.filter(sheet_category__icontains='Finish', inward__gt=0)
    elif report_type == 'outward':
        entries = entries.filter(dispatch__gt=0)
    else:
        return []

    for entry in entries:
        # Determine which quantity column to extract
        if report_type in ['inward', 'production']:
            qty = float(entry.inward)
        else:
            qty = float(entry.dispatch)
            
        results.append({
            'center': entry.sheet.center.name,
            'date': entry.sheet.date.strftime('%d-%m-%Y'),
            'raw_date': entry.sheet.date.strftime('%Y-%m-%d'),
            'item': entry.master_item.name.upper(),
            'category': entry.sheet_category,
            'qty': qty,
            'image_url': entry.sheet.image.url if entry.sheet.image else None
        })
        
    return results

@require_app_access('inventory')
def sorting_report_view(request):
    classifications = CenterClassification.objects.all().order_by('name')
    return render(request, 'inventory/sorting_report.html', {'classifications': classifications})

@require_app_access('inventory')
def api_sorting_data(request):
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        report_type = request.GET.get('type', 'production')
        mode = request.GET.get('mode', 'individual')
        group = request.GET.get('group', 'ALL')
        
        if not start_date or not end_date:
            return JsonResponse({'status': 'error', 'message': 'Date range required'})
            
        if mode == 'summary':
            centers = Center.objects.all()
            if group == 'HD': centers = centers.filter(category='HD')
            elif group == 'LD': centers = centers.filter(category='LD')
            elif group.isdigit(): centers = centers.filter(classification_id=group)

            center_ids = list(centers.values_list('id', flat=True))
            
            cmis = CenterMasterItem.objects.filter(center_id__in=center_ids)
            order_map = {(cmi.center_id, cmi.name.upper()): cmi.display_order for cmi in cmis}

            entries = StockEntry.objects.filter(
                sheet__date__range=[start_date, end_date],
                sheet__center_id__in=center_ids
            ).select_related('sheet__center', 'master_item').order_by('sheet__date')

            summary = {}
            for e in entries:
                c_id = e.sheet.center_id
                c_name = e.sheet.center.name
                i_name = e.master_item.name.upper()
                cat = e.sheet_category
                date = e.sheet.date

                if c_name not in summary:
                    summary[c_name] = {'raw': {}, 'fin': {}}
                
                cat_key = 'raw' if 'Raw' in cat else 'fin'
                key = i_name

                if key not in summary[c_name][cat_key]:
                    order = order_map.get((c_id, i_name), 9999)
                    summary[c_name][cat_key][key] = {
                        'item': i_name,
                        'order': order,
                        'op': float(e.opening_balance),
                        'inw': 0.0, 'prod': 0.0, 'disp': 0.0,
                        'clo': float(e.closing_balance),
                        'first_date': date,
                        'last_date': date
                    }

                item_data = summary[c_name][cat_key][key]
                if 'Raw' in cat: item_data['inw'] += float(e.inward)
                else: item_data['prod'] += float(e.inward)
                item_data['disp'] += float(e.dispatch)

                if date < item_data['first_date']:
                    item_data['op'] = float(e.opening_balance)
                    item_data['first_date'] = date
                if date >= item_data['last_date']:
                    item_data['clo'] = float(e.closing_balance)
                    item_data['last_date'] = date

            results = {}
            for c_name, cats in summary.items():
                results[c_name] = {
                    'raw': sorted(cats['raw'].values(), key=lambda x: (x['order'], x['item'])),
                    'fin': sorted(cats['fin'].values(), key=lambda x: (x['order'], x['item']))
                }
        else:
            results = get_report_data(start_date, end_date, report_type)
        return JsonResponse({'status': 'success', 'data': results})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_app_access('inventory')
def download_sorting_excel(request):
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        report_type = request.GET.get('type', 'production')
        mode = request.GET.get('mode', 'individual')
        group = request.GET.get('group', 'ALL')
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        
        font_bold = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))
        
        wb = openpyxl.Workbook()
        
        if mode == 'summary':
            brand_fill = PatternFill(start_color="01444E", end_color="01444E", fill_type="solid")
            
            centers = Center.objects.all()
            if group == 'HD': centers = centers.filter(category='HD')
            elif group == 'LD': centers = centers.filter(category='LD')
            elif group.isdigit(): centers = centers.filter(classification_id=group)
            
            center_ids = list(centers.values_list('id', flat=True))
            
            cmis = CenterMasterItem.objects.filter(center_id__in=center_ids)
            order_map = {(cmi.center_id, cmi.name.upper()): cmi.display_order for cmi in cmis}

            entries = StockEntry.objects.filter(
                sheet__date__range=[start_date, end_date],
                sheet__center_id__in=center_ids
            ).select_related('sheet__center', 'master_item').order_by('sheet__date')
            
            summary = {}
            for e in entries:
                c_id = e.sheet.center_id
                c_name = e.sheet.center.name
                i_name = e.master_item.name.upper()
                cat = e.sheet_category
                date = e.sheet.date
                
                if c_name not in summary:
                    summary[c_name] = {'raw': {}, 'fin': {}}
                
                cat_key = 'raw' if 'Raw' in cat else 'fin'
                key = i_name

                if key not in summary[c_name][cat_key]:
                    order = order_map.get((c_id, i_name), 9999)
                    summary[c_name][cat_key][key] = {
                        'item': i_name, 'order': order,
                        'op': float(e.opening_balance), 'inw': 0.0, 'prod': 0.0, 'disp': 0.0,
                        'clo': float(e.closing_balance), 'first_date': date, 'last_date': date
                    }
                
                item_data = summary[c_name][cat_key][key]
                if 'Raw' in cat: item_data['inw'] += float(e.inward)
                else: item_data['prod'] += float(e.inward)
                item_data['disp'] += float(e.dispatch)
                
                if date < item_data['first_date']:
                    item_data['op'] = float(e.opening_balance)
                    item_data['first_date'] = date
                if date >= item_data['last_date']:
                    item_data['clo'] = float(e.closing_balance)
                    item_data['last_date'] = date
            
            start_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_obj = datetime.strptime(end_date, '%Y-%m-%d')
            start_str = start_obj.strftime('%d %b %Y')
            end_str = end_obj.strftime('%d %b %Y')
            
            wb.remove(wb.active)
            if not summary:
                ws = wb.create_sheet("No Data")
                ws.cell(row=1, column=1, value="No data found for selected range and group.")
            else:
                for c_name, cats in sorted(summary.items()):
                    safe_title = c_name[:31].replace('/', '-').replace('\\', '-')
                    ws = wb.create_sheet(safe_title)
                    
                    raw_list = sorted(cats['raw'].values(), key=lambda x: (x['order'], x['item']))
                    fin_list = sorted(cats['fin'].values(), key=lambda x: (x['order'], x['item']))

                    current_row = 1
                    
                    if raw_list:
                        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                        title_cell = ws.cell(row=current_row, column=1, value="RAW MATERIALS")
                        title_cell.font = Font(bold=True, size=14, color="01444E")
                        current_row += 1
                        
                        headers = ["Item Name", f"Opening on {start_str}", "Total Inward", "Total dispatch/used for production", f"Closing on {end_str}"]
                        for col_num, header in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col_num, value=header)
                            cell.font = font_bold
                            cell.fill = brand_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border
                        current_row += 1
                        
                        for item in raw_list:
                            ws.cell(row=current_row, column=1, value=item['item']).border = thin_border
                            ws.cell(row=current_row, column=2, value=item['op']).border = thin_border
                            ws.cell(row=current_row, column=3, value=item['inw']).border = thin_border
                            ws.cell(row=current_row, column=4, value=item['disp']).border = thin_border
                            ws.cell(row=current_row, column=5, value=item['clo']).border = thin_border
                            current_row += 1
                        
                        current_row += 2
                        
                    if fin_list:
                        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                        title_cell = ws.cell(row=current_row, column=1, value="FINISHED GOODS")
                        title_cell.font = Font(bold=True, size=14, color="01444E")
                        current_row += 1
                        
                        headers = ["Item Name", f"Opening on {start_str}", "Total Inward/Production", "Total Dispatch", f"Closing on {end_str}"]
                        for col_num, header in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col_num, value=header)
                            cell.font = font_bold
                            cell.fill = brand_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border
                        current_row += 1
                        
                        for item in fin_list:
                            ws.cell(row=current_row, column=1, value=item['item']).border = thin_border
                            ws.cell(row=current_row, column=2, value=item['op']).border = thin_border
                            ws.cell(row=current_row, column=3, value=item['prod']).border = thin_border
                            ws.cell(row=current_row, column=4, value=item['disp']).border = thin_border
                            ws.cell(row=current_row, column=5, value=item['clo']).border = thin_border
                            current_row += 1
                    
                    for col_idx, col in enumerate(ws.columns, 1):
                        max_length = 0
                        column = get_column_letter(col_idx)
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                            except: pass
                        ws.column_dimensions[column].width = max_length + 2
            
            start_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_obj = datetime.strptime(end_date, '%Y-%m-%d')
            group_label = group if group in ['ALL', 'HD', 'LD'] else "Group"
            if group.isdigit():
                cls = CenterClassification.objects.filter(id=group).first()
                if cls: group_label = cls.name
            
            if group_label == 'ALL': group_label = 'All Centers'
            elif group_label == 'HD': group_label = 'HD Centers'
            elif group_label == 'LD': group_label = 'LD Centers'
            
            filename = f"Stock Report {group_label} {start_obj.strftime('%d %b')} to {end_obj.strftime('%d %b %Y')}.xlsx"
            
        else:
            results = get_report_data(start_date, end_date, report_type)
            ws = wb.active
            ws.title = f"{report_type.capitalize()}_Report"
            
            header_fill = PatternFill(start_color="FF8965", end_color="FF8965", fill_type="solid")
            qty_label = "Dispatch Qty" if report_type == 'outward' else ("Production Qty" if report_type == 'production' else "Inward Qty")
            
            headers = ["Center Name", "Date", "Item Name", "Category", qty_label]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = font_bold
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
            for row_num, row_data in enumerate(results, 2):
                ws.cell(row=row_num, column=1, value=row_data['center']).border = thin_border
                ws.cell(row=row_num, column=2, value=row_data['date']).border = thin_border
                ws.cell(row=row_num, column=3, value=row_data['item']).border = thin_border
                ws.cell(row=row_num, column=4, value=row_data['category']).border = thin_border
                ws.cell(row=row_num, column=5, value=row_data['qty']).border = thin_border
                
            for col_idx, col in enumerate(ws.columns, 1):
                max_length = 0
                column = get_column_letter(col_idx)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = max_length + 2
                
            filename = f"{report_type.capitalize()}_Report_{start_date}_to_{end_date}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)

# ==========================================
# 14. DATA SYNC & BACKUP HUB (SUPERUSER ONLY)
# ==========================================

@require_app_access('inventory')
def data_sync_hub(request):
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Superuser access required.")
    return render(request, 'inventory/sync_hub.html')

@require_app_access('inventory')
def export_full_system(request):
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Superuser access required.")

    db_out = StringIO()
    call_command('dumpdata', exclude=['contenttypes', 'auth.Permission', 'sessions', 'admin.LogEntry'], stdout=db_out)

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('database_backup.json', db_out.getvalue())
        
        media_root = settings.MEDIA_ROOT
        if os.path.exists(media_root):
            for root, dirs, files in os.walk(media_root):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, media_root)
                    zf.write(file_path, os.path.join('media', arcname))
    
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    
    date_str = datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Nexus_Full_Backup_{date_str}.zip"'
    return response

@require_app_access('inventory')
@require_POST
def import_full_system(request):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Superuser access required.'}, status=403)

    try:
        zip_file = request.FILES.get('backup_file')
        if not zip_file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'})
            
        with zipfile.ZipFile(zip_file, 'r') as zf:
            if 'database_backup.json' in zf.namelist():
                with tempfile.NamedTemporaryFile(delete=False, suffix='.json') as tmp:
                    tmp.write(zf.read('database_backup.json'))
                    tmp_path = tmp.name
                
                try:
                    call_command('loaddata', tmp_path)
                finally:
                    os.remove(tmp_path) 
            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid backup file: Missing database JSON.'})
                    
            media_root = settings.MEDIA_ROOT
            for item in zf.namelist():
                if item.startswith('media/'):
                    rel_path = item[6:] 
                    if rel_path: 
                        target_path = os.path.join(media_root, rel_path)
                        os.makedirs(os.path.dirname(target_path), exist_ok=True)
                        with open(target_path, 'wb') as f:
                            f.write(zf.read(item))
                            
        return JsonResponse({'status': 'success', 'message': 'System successfully restored!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# ==========================================
# 15. CENTER MASTER TEMPLATE BUILDER
# ==========================================

@require_feature_access('can_edit_center_templates')
def center_template_builder(request, center_id):
    center = get_object_or_404(Center, id=center_id)
    items = center.master_items.all().order_by('display_order')
    
    items_list = list(items.values('id', 'name', 'category', 'display_order'))
    
    return render(request, 'inventory/center_template_builder.html', {
        'center': center,
        'items_json': json.dumps(items_list)
    })

@require_feature_access('can_edit_center_templates')
@require_POST
def api_save_center_template(request, center_id):
    try:
        center = get_object_or_404(Center, id=center_id)
        data = json.loads(request.body)
        items = data.get('items', [])
        
        received_ids = []
        with transaction.atomic():
            for idx, item in enumerate(items):
                item_id = item.get('id')
                name = item.get('name', '').strip().upper() 
                cat = item.get('category', 'Raw Material')
                
                if not name: continue
                
                if item_id and str(item_id).isdigit():
                    obj = CenterMasterItem.objects.filter(id=item_id, center=center).first()
                    if obj:
                        obj.name = name
                        obj.category = cat
                        obj.display_order = idx
                        obj.save()
                        received_ids.append(obj.id)
                    else:
                        obj = CenterMasterItem.objects.create(center=center, name=name, category=cat, display_order=idx)
                        received_ids.append(obj.id)
                else:
                    obj = CenterMasterItem.objects.create(center=center, name=name, category=cat, display_order=idx)
                    received_ids.append(obj.id)
                    
            CenterMasterItem.objects.filter(center=center).exclude(id__in=received_ids).delete()
            
        return JsonResponse({'status': 'success', 'message': 'Master Template saved successfully!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# ==========================================
# 16. ALIAS MANAGEMENT ENGINE
# ==========================================

@require_feature_access('can_edit_center_templates')
def center_aliases(request, center_id):
    center = get_object_or_404(Center, id=center_id)
    
    masters = center.master_items.prefetch_related('centeritemalias_set').order_by('name')
    
    master_data = []
    for m in masters:
        master_data.append({
            'id': m.id,
            'name': m.name,
            'category': m.category,
            'aliases': [{'id': a.id, 'name': a.scanned_name} for a in m.centeritemalias_set.all()]
        })
        
    return render(request, 'inventory/center_aliases.html', {
        'center': center,
        'masters_json': json.dumps(master_data) 
    })

@require_feature_access('can_edit_center_templates')
@require_POST
def api_add_center_alias(request, center_id):
    try:
        data = json.loads(request.body)
        master_id = data.get('master_id')
        alias_name = data.get('alias_name', '').strip().upper()

        if not master_id or not alias_name:
            return JsonResponse({'status': 'error', 'message': 'Missing data'})

        master = get_object_or_404(CenterMasterItem, id=master_id, center_id=center_id)
        
        if CenterItemAlias.objects.filter(center_id=center_id, scanned_name__iexact=alias_name).exists():
            return JsonResponse({'status': 'error', 'message': f'"{alias_name}" is already mapped in this center!'})

        alias = CenterItemAlias.objects.create(center_id=center_id, mapped_to=master, scanned_name=alias_name)
        
        return JsonResponse({'status': 'success', 'alias_id': alias.id, 'alias_name': alias.scanned_name})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_feature_access('can_edit_center_templates')
@require_POST
def api_delete_alias(request, alias_id):
    try:
        alias = get_object_or_404(CenterItemAlias, id=alias_id)
        alias_name = alias.scanned_name
        alias.delete()
        return JsonResponse({'status': 'success', 'message': f'Alias "{alias_name}" deleted.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# ==========================================
# 17. TEMPLATE BUILDER (PRINT SHEETS)
# ==========================================

@require_app_access('inventory')
def template_builder_list(request):
    templates = PrintTemplate.objects.annotate(item_count=Count('items')).order_by('-created_at')
    return render(request, 'inventory/template_list.html', {'templates': templates})

@require_feature_access('can_edit_center_templates')
def template_builder_create(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            template_name = data.get('template_name', '').strip()
            center_ids = data.get('center_ids', [])
            template_ids = data.get('template_ids', [])

            if not template_name:
                return JsonResponse({'status': 'error', 'message': 'Template name is required.'})
            if PrintTemplate.objects.filter(name__iexact=template_name).exists():
                return JsonResponse({'status': 'error', 'message': 'Template with this name already exists.'})

            unique_items = {} 

            if center_ids:
                center_items = CenterMasterItem.objects.filter(center_id__in=center_ids).order_by('center_id', 'display_order')
                for ci in center_items:
                    key = (ci.name.upper(), ci.category)
                    if key not in unique_items:
                        unique_items[key] = ci.display_order 
            
            if template_ids:
                temp_items = PrintTemplateItem.objects.filter(template_id__in=template_ids).order_by('template_id', 'display_order')
                for ti in temp_items:
                    key = (ti.name.upper(), ti.category)
                    if key not in unique_items:
                        unique_items[key] = ti.display_order

            if not unique_items:
                return JsonResponse({'status': 'error', 'message': 'No items found in selected sources.'})

            with transaction.atomic():
                new_template = PrintTemplate.objects.create(name=template_name)
                
                if center_ids:
                    new_template.source_centers.set(center_ids)
                if template_ids:
                    for t_id in template_ids:
                        parent_template = PrintTemplate.objects.get(id=t_id)
                        new_template.source_centers.add(*parent_template.source_centers.all())
                
                items_to_create = []
                
                sorted_items = sorted(unique_items.items(), key=lambda x: (x[1], x[0][0]))
                
                for idx, ((name, cat), original_order) in enumerate(sorted_items):
                    items_to_create.append(PrintTemplateItem(
                        template=new_template, name=name, category=cat, display_order=idx
                    ))
                
                PrintTemplateItem.objects.bulk_create(items_to_create)

            return JsonResponse({'status': 'success', 'url': f'/inventory/templates/{new_template.id}/'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    centers = Center.objects.all().order_by('name')
    templates = PrintTemplate.objects.all().order_by('name')
    
    centers_data = [{'id': c.id, 'name': c.name, 'category': c.category} for c in centers]
    templates_data = [{'id': t.id, 'name': t.name} for t in templates]

    return render(request, 'inventory/template_create.html', {
        'centers_json': json.dumps(centers_data),
        'templates_json': json.dumps(templates_data)
    })

@require_app_access('inventory')
def template_builder_view(request, template_id):
    template = get_object_or_404(PrintTemplate, id=template_id)
    raw_items = template.items.filter(category='Raw Material').order_by('display_order', 'name')
    finished_items = template.items.filter(category='Finished Goods').order_by('display_order', 'name')
    
    return render(request, 'inventory/template_view.html', {
        'template': template,
        'raw_items': raw_items,
        'finished_items': finished_items
    })
    
@require_feature_access('can_edit_center_templates')
@require_POST
def api_delete_template(request, template_id):
    try:
        template = get_object_or_404(PrintTemplate, id=template_id)
        template.delete()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_app_access('inventory')
def template_item_vs_center_download(request, template_id):
    template = get_object_or_404(PrintTemplate, id=template_id)
    items = template.items.all().order_by('-category', 'name')
    source_centers = template.source_centers.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standardization_Matrix"
    
    font_bold = Font(bold=True)
    
    headers = ["Centers List", "Type", "Item Name", "Standardized Name"]
    for col_num, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_num, value=h)
        c.font = font_bold
        c.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        
    row_num = 2
    for item in items:
        if source_centers.exists():
            centers = CenterMasterItem.objects.filter(name__iexact=item.name, center__in=source_centers).select_related('center')
        else:
            centers = CenterMasterItem.objects.filter(name__iexact=item.name).select_related('center')
            
        center_names = ", ".join(sorted(set([c.center.name for c in centers])))
        if not center_names:
            center_names = "-"
            
        type_badge = "RM" if item.category == 'Raw Material' else "FG"
        
        ws.cell(row=row_num, column=1, value=center_names)
        ws.cell(row=row_num, column=2, value=type_badge).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=3, value=item.name)
        ws.cell(row=row_num, column=4, value="") 
        
        row_num += 1
        
    ws.column_dimensions['A'].width = 15  
    ws.column_dimensions['B'].width = 10  
    ws.column_dimensions['C'].width = 35  
    ws.column_dimensions['D'].width = 35  
    
    ws.freeze_panes = 'A2'
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{template.name}_Standardization_Matrix.xlsx"'
    wb.save(response)
    return response

@require_app_access('inventory')
@require_POST
def template_builder_download(request, template_id):
    template = get_object_or_404(PrintTemplate, id=template_id)
    
    data = json.loads(request.POST.get('payload', '{}'))
    print_title = data.get('print_title', '').upper()
    raw_cols = data.get('raw_columns', ['Opening', 'Inward', 'Dispatch', 'Closing'])
    fin_cols = data.get('finished_columns', ['Opening', 'Production', 'Dispatch', 'Closing'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Print_Template"

    font_bold = Font(name="Calibri", size=12, bold=True)
    font_normal = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )

    ws.sheet_format.defaultRowHeight = 25

    raw_items = template.items.filter(category='Raw Material').order_by('display_order', 'name')
    fin_items = template.items.filter(category='Finished Goods').order_by('display_order', 'name')
    
    max_name_length = 15
    for item in list(raw_items) + list(fin_items):
        if len(item.name) > max_name_length:
            max_name_length = len(item.name)

    max_total_cols = max(len(raw_cols), len(fin_cols)) + 2
    if max_total_cols < 4: max_total_cols = 4

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_total_cols-2)
    ws.cell(row=1, column=1, value=print_title).font = title_font
    ws.cell(row=1, column=1).alignment = align_left

    ws.cell(row=1, column=max_total_cols-1, value="DATE:").font = font_bold
    ws.cell(row=1, column=max_total_cols-1).alignment = align_right
    ws.cell(row=1, column=max_total_cols, value="").alignment = align_left

    current_row = 3

    def draw_section(title, items, columns):
        nonlocal current_row
        
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns)+2)
        cell = ws.cell(row=current_row, column=1, value=title.upper())
        cell.font = Font(name="Calibri", size=14, bold=True)
        cell.alignment = align_center
        for col in range(1, len(columns)+3):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        headers = ["SR NO.", "ITEM NAME"] + [c.upper() for c in columns]
        for col_num, h in enumerate(headers, 1):
            c = ws.cell(row=current_row, column=col_num, value=h)
            c.font = font_bold
            c.alignment = align_center
            c.border = thin_border
        current_row += 1

        for i, item in enumerate(items, 1):
            ws.cell(row=current_row, column=1, value=i).alignment = align_center
            ws.cell(row=current_row, column=2, value=item.name).alignment = align_left
            
            for col in range(1, len(headers)+1):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).font = font_normal
            current_row += 1
            
        current_row += 2 

    if raw_items.exists():
        draw_section("RAW MATERIAL", raw_items, raw_cols)

    if fin_items.exists():
        draw_section("FINISHED GOODS", fin_items, fin_cols)

    ws.column_dimensions['A'].width = 8   
    ws.column_dimensions['B'].width = max_name_length + 3  
    
    dynamic_width = 50 / max_total_cols if max_total_cols > 0 else 15
    dynamic_width = max(12, dynamic_width) 

    for col in range(3, max_total_cols + 3):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = dynamic_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{template.name}_Print_Sheet.xlsx"'
    wb.save(response)
    return response

# ==================================================
# 18. CUSTOM DYNAMIC REPORTS (CENTER VS ITEM PIVOT)
# ==================================================

@require_app_access('inventory')
def custom_report_page(request):
    centers = Center.objects.all().order_by('name')
    centers_list = list(centers.values('id', 'name')) 
    
    return render(request, 'inventory/custom_report.html', {
        'centers': centers,          
        'centers_list': centers_list  
    })
    
@require_app_access('inventory')
def api_get_report_items(request):
    date_str = request.GET.get('date')
    center_ids = request.GET.getlist('centers[]') 
    
    if not date_str:
        return JsonResponse({'status': 'error', 'message': 'Date is required.'})

    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        sheets = StockSheet.objects.filter(date=date_obj)
        
        if center_ids and center_ids[0] != '':
            sheets = sheets.filter(center_id__in=center_ids)
            
        items = StockEntry.objects.filter(sheet__in=sheets).values_list('master_item__name', flat=True).distinct().order_by('master_item__name')
        
        return JsonResponse({'status': 'success', 'items': list(items)})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_app_access('inventory')
@require_POST
def api_generate_custom_report(request):
    try:
        data = json.loads(request.body)
        date_str = data.get('date')
        center_ids = data.get('centers', [])
        item_names = data.get('items', [])
        
        if not date_str:
            return JsonResponse({'status': 'error', 'message': 'Date is required.'})

        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        sheets = StockSheet.objects.filter(date=date_obj).select_related('center')
        if center_ids:
            sheets = sheets.filter(center_id__in=center_ids)

        valid_centers = list(sheets.values_list('center__name', flat=True).distinct().order_by('center__name'))
        
        entries = StockEntry.objects.filter(sheet__in=sheets).select_related('sheet__center', 'master_item')
        if item_names:
            entries = entries.filter(master_item__name__in=item_names)

        report_data = {}
        category_totals = {}
        grand_totals = {c: 0 for c in valid_centers}
        grand_totals['Total'] = 0

        for entry in entries:
            item = entry.master_item.name.upper()
            center = entry.sheet.center.name
            closing = float(entry.closing_balance)
            category = entry.sheet_category or "Uncategorized"
            
            if item not in report_data:
                report_data[item] = {'Category': category, 'Item': item, 'Total': 0}
                for c in valid_centers:
                    report_data[item][c] = 0
            
            report_data[item][center] += closing
            report_data[item]['Total'] += closing

            if category not in category_totals:
                category_totals[category] = {'Total': 0}
                for c in valid_centers:
                    category_totals[category][c] = 0
            
            category_totals[category][center] += closing
            category_totals[category]['Total'] += closing

            grand_totals[center] += closing
            grand_totals['Total'] += closing

        # --- REORDERING LOGIC: Totals at the top ---
        final_data = []

        def add_summary_row(cat_key, type_label, item_label, is_grand=False):
            if cat_key in category_totals or is_grand:
                totals = grand_totals if is_grand else category_totals[cat_key]
                row = {'Category': type_label, 'Item': item_label, 'Total': totals['Total'], 'is_summary': True}
                for c in valid_centers:
                    row[c] = totals.get(c, 0)
                final_data.append(row)

        # 1. Pinned Summary Rows
        add_summary_row('Raw Material', 'RM', 'RAW MATERIAL TOTAL')
        add_summary_row('Finished Goods', 'FG', 'FINISHED GOODS TOTAL')
        add_summary_row('ALL', 'ALL', 'GRAND TOTAL', is_grand=True)

        # 2. Individual Item Rows (Sorted FG first, then RM)
        sorted_items = sorted(list(report_data.values()), key=lambda x: (x['Category'] == 'Raw Material', x['Item']))
        for row in sorted_items:
            # Map long category names to short codes for the table
            if row['Category'] == 'Raw Material': row['Category'] = 'RM'
            elif row['Category'] == 'Finished Goods': row['Category'] = 'FG'
            final_data.append(row)

        return JsonResponse({'status': 'success', 'columns': valid_centers, 'data': final_data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@require_app_access('inventory')
def export_custom_report_excel(request):
    date_str = request.GET.get('date')
    center_ids = request.GET.get('centers', '').split(',') if request.GET.get('centers') else []
    item_names = request.GET.get('items', '').split(',') if request.GET.get('items') else []
    
    center_ids = [c for c in center_ids if c]
    item_names = [i for i in item_names if i]

    if not date_str: return HttpResponse("Date is required", status=400)

    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    sheets = StockSheet.objects.filter(date=date_obj).select_related('center')
    if center_ids: sheets = sheets.filter(center_id__in=center_ids)

    valid_centers = list(sheets.values_list('center__name', flat=True).distinct().order_by('center__name'))
    
    entries = StockEntry.objects.filter(sheet__in=sheets).select_related('sheet__center', 'master_item')
    if item_names: entries = entries.filter(master_item__name__in=item_names)

    report_data = {}
    category_totals = {}
    grand_totals = {c: 0 for c in valid_centers}
    grand_totals['Total'] = 0

    for entry in entries:
        item = entry.master_item.name.upper()
        center = entry.sheet.center.name
        closing = float(entry.closing_balance)
        category = entry.sheet_category or "Uncategorized"
        
        if item not in report_data:
            report_data[item] = {'Category': category, 'Item': item, 'Total': 0}
            for c in valid_centers: report_data[item][c] = 0
                
        report_data[item][center] += closing
        report_data[item]['Total'] += closing

        if category not in category_totals:
            category_totals[category] = {'Total': 0}
            for c in valid_centers: category_totals[category][c] = 0
        
        category_totals[category][center] += closing
        category_totals[category]['Total'] += closing

        grand_totals[center] += closing
        grand_totals['Total'] += closing

    # Build Final Array
    final_data = []
    def add_summary_row(cat_key, type_label, item_label, is_grand=False):
        if cat_key in category_totals or is_grand:
            totals = grand_totals if is_grand else category_totals[cat_key]
            row = {'Category': type_label, 'Item': item_label, 'Total': totals['Total'], 'is_summary': True}
            for c in valid_centers: row[c] = totals.get(c, 0)
            final_data.append(row)

    add_summary_row('Raw Material', 'RM', 'RAW MATERIAL TOTAL')
    add_summary_row('Finished Goods', 'FG', 'FINISHED GOODS TOTAL')
    add_summary_row('ALL', 'ALL', 'GRAND TOTAL', is_grand=True)

    sorted_items = sorted(list(report_data.values()), key=lambda x: (x['Category'] == 'Raw Material', x['Item']))
    for row in sorted_items:
        if row['Category'] == 'Raw Material': row['Category'] = 'RM'
        elif row['Category'] == 'Finished Goods': row['Category'] = 'FG'
        final_data.append(row)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Custom Closing Report"

    # Strict column order
    headers = ['Type', 'Material Name', 'Total'] + valid_centers
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))
    summary_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    for row_idx, row_dict in enumerate(final_data, 2):
        row_data = [row_dict['Category'], row_dict['Item'], row_dict['Total']]
        for c in valid_centers:
            val = row_dict.get(c, 0)
            # Replace 0 with '-' for the Excel file
            row_data.append('-' if val == 0 else val)
        
        ws.append(row_data)
        
        is_summary = row_dict.get('is_summary', False)
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            if col_num > 2: cell.alignment = Alignment(horizontal="center")
            if is_summary:
                cell.font = Font(bold=True)
                cell.fill = summary_fill

    ws.column_dimensions['A'].width = 10 
    ws.column_dimensions['B'].width = 35 
    ws.column_dimensions['C'].width = 15 
    for col_num in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from django.http import HttpResponse
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Custom_Closing_Report_{date_str}.xlsx'
    return response


# ==================================================
# 19. Item Aliases Export
# ==================================================
@require_app_access('inventory')
def export_center_aliases_excel(request):
    # 1. Fetch all aliases, pulling in the related center and mapped item to avoid N+1 queries
    aliases = CenterItemAlias.objects.select_related('center', 'mapped_to').all().order_by('center__name', 'mapped_to__name')

    # 2. Group the aliases
    # We want a structure like: {(Center Name, Master Name): [alias1, alias2, ...]}
    grouped_data = defaultdict(list)
    for alias in aliases:
        key = (alias.center.name, alias.mapped_to.name)
        grouped_data[key].append(alias.scanned_name)

    # 3. Build the Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alias Dictionary"

    # Headers
    headers = ['Center Name', 'Aliases (Scanned Names)', 'Actual Name in Nexus']
    ws.append(headers)

    # Style Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'), 
        right=Side(style='thin', color='E2E8F0'), 
        top=Side(style='thin', color='E2E8F0'), 
        bottom=Side(style='thin', color='E2E8F0')
    )

    # 4. Populate the Data
    row_idx = 2
    for (center_name, actual_name), alias_list in grouped_data.items():
        # Join the list of aliases into a single comma-separated string
        aliases_str = ", ".join(alias_list)
        
        ws.append([center_name, aliases_str, actual_name])
        
        # Apply borders and alignment to the data cells
        for col_num in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            if col_num == 1:
                cell.alignment = Alignment(horizontal="left", vertical="top")
            elif col_num == 2:
                # Wrap text for the aliases column so it doesn't bleed endlessly to the right
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True) 
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")
                
        row_idx += 1

    # 5. Adjust Column Widths
    ws.column_dimensions['A'].width = 25  # Center Name
    ws.column_dimensions['B'].width = 60  # Aliases (Needs more space)
    ws.column_dimensions['C'].width = 35  # Actual Name

    # 6. Return as a downloadable response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Nexus_Alias_Dictionary.xlsx"'
    wb.save(response)
    
    return response

@require_app_access('inventory')
def target_report_page(request):
    centers = list(Center.objects.values('id', 'name').order_by('name'))
    
    center_items = {}
    for cmi in CenterMasterItem.objects.filter(category='Raw Material').select_related('center'):
        if cmi.center_id not in center_items:
            center_items[cmi.center_id] = []
        center_items[cmi.center_id].append(cmi.name.upper())

    reports = TargetReportConfig.objects.prefetch_related('columns', 'columns__centers').all().order_by('name')
    saved_reports = []
    for r in reports:
        cols = []
        for c in r.columns.all():
            cols.append({
                # CHANGED: We now pull a list of center IDs
                'center_ids': list(c.centers.values_list('id', flat=True)), 
                'name': c.column_name,
                'target': float(c.monthly_target),
                'items': c.selected_items
            })
        saved_reports.append({'id': r.id, 'name': r.name, 'columns': cols})

    return render(request, 'inventory/target_report.html', {
        'centers_json': json.dumps(centers),
        'center_items_json': json.dumps(center_items),
        'saved_reports_json': json.dumps(saved_reports)
    })

@require_app_access('inventory')
@require_POST
def api_save_target_report(request):
    try:
        data = json.loads(request.body)
        report_id = data.get('report_id')
        name = data.get('name')
        columns = data.get('columns', [])

        if not name: return JsonResponse({'status': 'error', 'message': 'Report name required.'})
        if not columns: return JsonResponse({'status': 'error', 'message': 'Add at least one column.'})

        with transaction.atomic():
            if report_id:
                report = TargetReportConfig.objects.get(id=report_id)
                report.name = name
                report.save()
                report.columns.all().delete()
            else:
                report = TargetReportConfig.objects.create(name=name)

            for idx, col in enumerate(columns):
                # Clean up any empty selections from the UI
                valid_center_ids = [cid for cid in col.get('center_ids', []) if cid]
                if not valid_center_ids: continue 

                new_col = TargetReportColumn.objects.create(
                    report=report,
                    column_name=col['name'],
                    monthly_target=col['target'],
                    selected_items=col['items'],
                    display_order=idx
                )
                # Set the many-to-many relationship
                new_col.centers.set(valid_center_ids)

        return JsonResponse({'status': 'success', 'report_id': report.id, 'message': 'Saved!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def calculate_target_report_data(report_id, date_str):
    report = get_object_or_404(TargetReportConfig, id=report_id)
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    year, month = target_date.year, target_date.month
    _, days_in_month = calendar.monthrange(year, month)
    calc_days = target_date.day

    columns = report.columns.prefetch_related('centers').all()
    col_names = [c.column_name for c in columns]
    
    row_a, row_b, row_c = [], [], []
    col_data_map = []

    for c in columns:
        m_target = float(c.monthly_target)
        d_target = m_target / days_in_month if days_in_month else 0
        t_till_date = d_target * calc_days
        
        row_a.append(m_target)
        row_b.append(round(d_target, 2))
        row_c.append(round(t_till_date, 2))
        # CHANGED: Map now stores a list of multiple center IDs
        col_data_map.append({'center_ids': list(c.centers.values_list('id', flat=True)), 'items': c.selected_items})

    daily_matrix = {day: [0] * len(columns) for day in range(1, calc_days + 1)}
    
    entries = StockEntry.objects.filter(
        sheet__date__year=year,
        sheet__date__month=month,
        sheet__date__lte=target_date
    ).select_related('sheet', 'master_item')

    for entry in entries:
        day = entry.sheet.date.day
        c_id = entry.sheet.center_id
        item_name = entry.master_item.name.upper()
        inward_val = float(entry.inward)

        if inward_val == 0: continue

        for col_idx, col_config in enumerate(col_data_map):
            # CHANGED: Now checks if the transaction's center is IN the list of approved centers for this column
            if c_id in col_config['center_ids'] and item_name in [i.upper() for i in col_config['items']]:
                daily_matrix[day][col_idx] += inward_val

    daily_rows = []
    row_e = [0] * len(columns)
    
    for day in range(1, calc_days + 1):
        row_data = daily_matrix[day]
        for idx, val in enumerate(row_data): row_e[idx] += val
        daily_rows.append({
            'date': f"{day:02d}-{month:02d}-{year}",
            'data': row_data,
            'total': sum(row_data)
        })

    row_f = [round(row_c[i] - row_e[i], 2) for i in range(len(columns))]

    return {
        'report_name': report.name, 'month_name': f"{calendar.month_name[month]} {year}",
        'col_names': col_names, 'row_a': row_a, 'row_b': row_b, 'row_c': row_c,
        'daily_rows': daily_rows, 'row_e': row_e, 'row_f': row_f,
        'total_a': sum(row_a), 'total_b': sum(row_b), 'total_c': sum(row_c),
        'total_e': sum(row_e), 'total_f': sum(row_f)
    }
    
@require_app_access('inventory')
def api_run_target_report(request):
    report_id = request.GET.get('report_id')
    date_str = request.GET.get('date')
    try:
        data = calculate_target_report_data(report_id, date_str)
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_app_access('inventory')
def export_target_report_excel(request, report_id, date_str):
    try:
        data = calculate_target_report_data(report_id, date_str)
        
        # Re-parse dates to get days in month and calc days for the Excel formulas
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        year, month = target_date.year, target_date.month
        _, days_in_month = calendar.monthrange(year, month)
        calc_days = target_date.day

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Target Report"

        header_font = Font(bold=True)
        light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") 
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['S.R No', 'Particulars', 'Total'] + data['col_names']
        last_col_letter = get_column_letter(len(headers))

        # UPDATED: Added custom_format parameter here
        def style_row(row_idx, is_header=False, fill=None, custom_format=None):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx > 2 and (type(cell.value) in [int, float] or str(cell.value).startswith('=')):
                    # Use custom_format if provided, else use default comma format
                    cell.number_format = custom_format if custom_format else '#,##0'
                if is_header or fill:
                    cell.font = Font(bold=True)
                if fill:
                    cell.fill = fill

        # Row 1: Top Header
        ws.append([data['month_name'] + ' Month'] + [''] * (len(headers) - 1)) 
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.cell(row=1, column=1).fill = light_blue_fill
        ws.cell(row=1, column=1).font = Font(bold=True)

        # Row 2: Headers
        ws.append(headers)
        style_row(2, is_header=True, fill=light_blue_fill)

        # Row 3: Monthly Target
        row_3 = ['A', 'Monthly Purchase Target', f'=SUM(D3:{last_col_letter}3)'] + data['row_a']
        ws.append(row_3)
        style_row(3, is_header=True)

        # Row 4: Daily Target
        row_4 = ['B', 'Daily Purchase Target', f'=SUM(D4:{last_col_letter}4)']
        for col_idx in range(4, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            row_4.append(f'={col_letter}3/{days_in_month}')
        ws.append(row_4)
        style_row(4)

        # Row 5: Target till Date
        row_5 = ['C', 'Target till Date', f'=SUM(D5:{last_col_letter}5)']
        for col_idx in range(4, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            row_5.append(f'={col_letter}4*{calc_days}')
        ws.append(row_5)
        style_row(5)

        ws.append([]) # Row 6 Spacer

        # Row 7: Daily Inward Header
        ws.append(['D', 'Daily Inward'])
        ws.cell(row=7, column=2).font = Font(bold=True)

        # Row 8+: Daily Rows
        current_row = 8
        for row in data['daily_rows']:
            r_data = ['', row['date'], f'=SUM(D{current_row}:{last_col_letter}{current_row})']
            for val in row['data']:
                r_data.append('-' if val == 0 else val)
            ws.append(r_data)
            style_row(current_row)
            current_row += 1

        last_daily_row = current_row - 1
        ws.append([]) # Spacer
        current_row += 1

        # Row E: Total Inward till date
        row_e_idx = current_row
        row_e = ['E', 'Total Inward till date', f'=SUM(D{row_e_idx}:{last_col_letter}{row_e_idx})']
        for col_idx in range(4, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            if last_daily_row >= 8:
                row_e.append(f'=SUM({col_letter}8:{col_letter}{last_daily_row})')
            else:
                row_e.append(0)
        ws.append(row_e)
        style_row(row_e_idx, is_header=True, fill=light_blue_fill)
        current_row += 1

        # Row F: Shortfall
        row_f_idx = current_row
        row_f = ['F', 'Shortfall Purchase (C - E)', f'=SUM(D{row_f_idx}:{last_col_letter}{row_f_idx})']
        for col_idx in range(4, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            row_f.append(f'={col_letter}5-{col_letter}{row_e_idx}')
        ws.append(row_f)
        
        # UPDATED: We pass the native Excel formatting string into style_row here!
        style_row(row_f_idx, is_header=True, fill=light_blue_fill, custom_format='#,##0;[Green](#,##0);0')

        # Widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        for i in range(4, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Format the date to DD MM YYYY for the filename
        formatted_filename_date = target_date.strftime('%d %m %Y')
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Wrapped in quotes to safely handle spaces in the filename
        response['Content-Disposition'] = f'attachment; filename="Target Report {formatted_filename_date}.xlsx"'
        return response

    except Exception as e:
        return HttpResponse(str(e), status=500)

@require_app_access('inventory')
@require_POST
def api_delete_target_report(request):
    try:
        data = json.loads(request.body)
        report_id = data.get('report_id')

        if not report_id:
            return JsonResponse({'status': 'error', 'message': 'Report ID is required for deletion.'})

        # Find the report and delete it. 
        # (Because of on_delete=models.CASCADE in our models, all associated columns are automatically deleted too).
        report = get_object_or_404(TargetReportConfig, id=report_id)
        report_name = report.name
        report.delete()

        return JsonResponse({'status': 'success', 'message': f'"{report_name}" has been deleted.'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_app_access('inventory')
def custom_reports_landing(request):
    """Renders the hub page for all custom reports."""
    return render(request, 'inventory/custom_reports_landing.html')